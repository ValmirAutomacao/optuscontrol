"""
Rotas de Exportação de Dados para Contabilidade
Gera planilhas Excel, arquivos SPED Fiscal e outros formatos.
"""
from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from typing import List, Optional
from datetime import date, datetime
from io import BytesIO
import json
from ...db.supabase_client import supabase

router = APIRouter(prefix="/export", tags=["Export"])


@router.get("/lancamentos")
async def export_lancamentos_excel(
    company_id: str = Query(...),
    start_date: str = Query(..., description="Data início YYYY-MM-DD"),
    end_date: str = Query(..., description="Data fim YYYY-MM-DD"),
    include_invoices: bool = True,
    include_expenses: bool = True
):
    """
    Exporta lançamentos no formato Excel para contabilidade.
    Formato: DATA, ENTRADA/SAIDA, EMPRESA, NF, DATA EMISSAO NF, DISCRIMINACAO, VALOR R$
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl não instalado")
    
    lancamentos = []
    
    # Buscar NF-e (entrada de mercadorias)
    if include_invoices:
        result = supabase.table("invoices").select("*").eq("company_id", company_id).gte("issue_date", start_date).lte("issue_date", end_date).execute()
        for inv in (result.data or []):
            lancamentos.append({
                "data": inv.get("created_at", "")[:10],
                "tipo": "E",  # Entrada
                "empresa": inv.get("supplier_name", ""),
                "nf": inv.get("number", ""),
                "data_emissao": inv.get("issue_date", ""),
                "discriminacao": f"NF-e {inv.get('number', '')} - {inv.get('supplier_name', '')}",
                "valor": float(inv.get("total_value", 0) or 0)
            })
    
    # Buscar Despesas (saída de caixa)
    if include_expenses:
        result = supabase.table("receipts").select("*").eq("company_id", company_id).gte("receipt_date", start_date).lte("receipt_date", end_date).execute()
        for exp in (result.data or []):
            lancamentos.append({
                "data": exp.get("created_at", "")[:10],
                "tipo": "S",  # Saída
                "empresa": exp.get("establishment_name", ""),
                "nf": "",  # Cupom não tem NF
                "data_emissao": exp.get("receipt_date", ""),
                "discriminacao": f"Despesa - {exp.get('establishment_name', 'N/I')}",
                "valor": float(exp.get("total_amount", 0) or 0)
            })
    
    # Buscar Contas a Pagar pagas
    result = supabase.table("payables").select("*").eq("company_id", company_id).eq("status", "paid").gte("payment_date", start_date).lte("payment_date", end_date).execute()
    for pay in (result.data or []):
        lancamentos.append({
            "data": pay.get("payment_date", ""),
            "tipo": "S",  # Saída
            "empresa": pay.get("supplier_name", ""),
            "nf": "",
            "data_emissao": pay.get("due_date", ""),
            "discriminacao": pay.get("description", ""),
            "valor": float(pay.get("amount", 0) or 0)
        })
    
    # Ordenar por data
    lancamentos.sort(key=lambda x: x.get("data", ""))
    
    # Criar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lançamentos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4361EE", end_color="4361EE", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    year = start_date[:4]
    ws.merge_cells('A1:G1')
    ws['A1'] = f"TABELA DE LANÇAMENTOS {year}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Cabeçalhos
    headers = ["DATA", "ENTRADA/SAIDA", "EMPRESA", "NF", "DATA EMISSAO NF", "DISCRIMINACAO", "VALOR R$"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    # Dados
    for row_idx, lanc in enumerate(lancamentos, 4):
        ws.cell(row=row_idx, column=1, value=lanc["data"]).border = border
        ws.cell(row=row_idx, column=2, value=lanc["tipo"]).border = border
        ws.cell(row=row_idx, column=3, value=lanc["empresa"]).border = border
        ws.cell(row=row_idx, column=4, value=lanc["nf"]).border = border
        ws.cell(row=row_idx, column=5, value=lanc["data_emissao"]).border = border
        ws.cell(row=row_idx, column=6, value=lanc["discriminacao"]).border = border
        valor_cell = ws.cell(row=row_idx, column=7, value=lanc["valor"])
        valor_cell.border = border
        valor_cell.number_format = 'R$ #,##0.00'
    
    # Ajustar larguras
    column_widths = [12, 15, 35, 15, 18, 45, 15]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Totais
    last_row = len(lancamentos) + 4
    ws.cell(row=last_row, column=6, value="TOTAL ENTRADAS:").font = Font(bold=True)
    ws.cell(row=last_row, column=7, value=sum(l["valor"] for l in lancamentos if l["tipo"] == "E")).number_format = 'R$ #,##0.00'
    
    ws.cell(row=last_row + 1, column=6, value="TOTAL SAÍDAS:").font = Font(bold=True)
    ws.cell(row=last_row + 1, column=7, value=sum(l["valor"] for l in lancamentos if l["tipo"] == "S")).number_format = 'R$ #,##0.00'
    
    # Salvar em buffer
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"lancamentos_{start_date}_{end_date}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/resumo")
async def get_export_summary(
    company_id: str = Query(...),
    start_date: str = Query(...),
    end_date: str = Query(...)
) -> dict:
    """Retorna resumo dos dados disponíveis para exportação."""
    invoices = supabase.table("invoices").select("id, total_value").eq("company_id", company_id).gte("issue_date", start_date).lte("issue_date", end_date).execute()
    expenses = supabase.table("receipts").select("id, total_amount").eq("company_id", company_id).gte("receipt_date", start_date).lte("receipt_date", end_date).execute()
    payables = supabase.table("payables").select("id, amount").eq("company_id", company_id).eq("status", "paid").gte("payment_date", start_date).lte("payment_date", end_date).execute()
    
    return {
        "period": {"start": start_date, "end": end_date},
        "invoices": {
            "count": len(invoices.data or []),
            "total": sum(float(i.get("total_value", 0) or 0) for i in (invoices.data or []))
        },
        "expenses": {
            "count": len(expenses.data or []),
            "total": sum(float(e.get("total_amount", 0) or 0) for e in (expenses.data or []))
        },
        "payables_paid": {
            "count": len(payables.data or []),
            "total": sum(float(p.get("amount", 0) or 0) for p in (payables.data or []))
        }
    }


@router.get("/sped-fiscal")
async def export_sped_fiscal(
    company_id: str = Query(...),
    start_date: str = Query(..., description="Data início YYYY-MM-DD"),
    end_date: str = Query(..., description="Data fim YYYY-MM-DD")
):
    """
    Exporta arquivo SPED Fiscal (EFD ICMS/IPI).
    Arquivo texto no padrão da Receita Federal.
    """
    from ...core.sped_generator import generate_sped_fiscal
    from io import StringIO
    
    # Buscar dados da empresa
    company_result = supabase.table("companies").select("*").eq("id", company_id).single().execute()
    company_data = company_result.data or {
        "name": "EMPRESA NÃO IDENTIFICADA",
        "cnpj": "",
        "state": "CE"
    }
    
    # Buscar NF-e
    invoices_result = supabase.table("invoices").select("*").eq("company_id", company_id).gte("issue_date", start_date).lte("issue_date", end_date).execute()
    invoices = invoices_result.data or []
    
    # Gerar arquivo SPED
    content = generate_sped_fiscal(
        company_data=company_data,
        start_date=start_date,
        end_date=end_date,
        invoices=invoices
    )
    
    output = BytesIO(content.encode('utf-8'))
    output.seek(0)
    
    filename = f"sped_fiscal_{start_date}_{end_date}.txt"
    
    return StreamingResponse(
        output,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/efd-contribuicoes")
async def export_efd_contribuicoes(
    company_id: str = Query(...),
    start_date: str = Query(..., description="Data início YYYY-MM-DD"),
    end_date: str = Query(..., description="Data fim YYYY-MM-DD")
):
    """
    Exporta arquivo EFD Contribuições (PIS/COFINS).
    Arquivo texto no padrão da Receita Federal.
    """
    from ...core.efd_contribuicoes_generator import generate_efd_contribuicoes
    
    # Buscar dados da empresa
    company_result = supabase.table("companies").select("*").eq("id", company_id).single().execute()
    company_data = company_result.data or {
        "name": "EMPRESA NÃO IDENTIFICADA",
        "cnpj": "",
        "state": "CE"
    }
    
    # Buscar NF-e
    invoices_result = supabase.table("invoices").select("*").eq("company_id", company_id).gte("issue_date", start_date).lte("issue_date", end_date).execute()
    invoices = invoices_result.data or []
    
    # Buscar Despesas
    expenses_result = supabase.table("receipts").select("*").eq("company_id", company_id).gte("receipt_date", start_date).lte("receipt_date", end_date).execute()
    expenses = expenses_result.data or []
    
    # Gerar arquivo EFD
    content = generate_efd_contribuicoes(
        company_data=company_data,
        start_date=start_date,
        end_date=end_date,
        invoices=invoices,
        expenses=expenses
    )
    
    output = BytesIO(content.encode('utf-8'))
    output.seek(0)
    
    filename = f"efd_contribuicoes_{start_date}_{end_date}.txt"
    
    return StreamingResponse(
        output,
        media_type="text/plain",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

